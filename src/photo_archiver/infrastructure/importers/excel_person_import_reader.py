"""Excel implementation of the person import reader port."""

from pathlib import Path
from typing import Any

from photo_archiver.application.dtos import PersonImportRow
from photo_archiver.application.ports import PersonImportReader


class ExcelPersonImportReader(PersonImportReader):
    """Read people from Excel workbook files."""

    def read(
        self,
        source_path: Path,
        *,
        has_header: bool = True,
        sheet_name: str | None = None,
    ) -> list[PersonImportRow]:
        """Read people from the selected worksheet in an Excel workbook."""
        from openpyxl import load_workbook

        path = Path(source_path)
        if not path.exists():
            raise FileNotFoundError(f"Person import file does not exist: {path}")
        if not path.is_file():
            raise IsADirectoryError(f"Person import path is not a file: {path}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
            rows: list[PersonImportRow] = []
            for row_number, cells in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = list(cells)
                if has_header and row_number == 1:
                    continue
                if self._is_blank_row(values):
                    continue
                rows.append(self._build_row(values, row_number))
            return rows
        finally:
            workbook.close()

    @staticmethod
    def _is_blank_row(values: list[Any]) -> bool:
        """Return whether all cell values are empty or whitespace."""
        return not any(str(value).strip() for value in values if value is not None)

    @classmethod
    def _build_row(cls, values: list[Any], row_number: int) -> PersonImportRow:
        """Build a normalized person import row from worksheet cell values."""
        padded_values = [cls._normalize_cell(value) for value in values[:4]]
        padded_values.extend([None] * (4 - len(padded_values)))
        name, identity, department, note = padded_values
        return PersonImportRow(
            name=name or "",
            identity=identity,
            department=department,
            note=note,
            row_number=row_number,
        )

    @staticmethod
    def _normalize_cell(value: Any) -> str | None:
        """Normalize an Excel cell value into an optional string field."""
        if value is None:
            return None
        normalized = str(value).strip()
        return normalized or None