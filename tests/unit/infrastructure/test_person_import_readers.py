"""Tests for person import reader infrastructure adapters."""

from importlib.util import find_spec
from pathlib import Path

import pytest

from photo_archiver.infrastructure import ExcelPersonImportReader, TxtPersonImportReader


requires_openpyxl = pytest.mark.skipif(find_spec("openpyxl") is None, reason="openpyxl is not installed")


def save_workbook(path: Path, rows: list[list[object | None]], *, title: str = "Sheet") -> None:
    """Create an Excel workbook with the provided rows."""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_txt_person_import_reader_reads_delimited_rows(tmp_path: Path) -> None:
    """Read normalized person rows from TXT/CSV-style sources."""
    source = tmp_path / "people.txt"
    source.write_text(
        "name,identity,department,note\n"
        "Alice,A001,Archive,Lead\n"
        "  \n"
        "Bob,B002,,\n",
        encoding="utf-8",
    )

    rows = TxtPersonImportReader().read(source)

    assert [row.name for row in rows] == ["Alice", "Bob"]
    assert rows[0].identity == "A001"
    assert rows[0].department == "Archive"
    assert rows[0].note == "Lead"
    assert rows[0].row_number == 2
    assert rows[1].department is None


@requires_openpyxl
def test_excel_person_import_reader_reads_headered_active_sheet(tmp_path: Path) -> None:
    """Read rows from the active worksheet while skipping a header row."""
    source = tmp_path / "people.xlsx"
    save_workbook(
        source,
        [
            ["name", "identity", "department", "note"],
            ["Alice", "A001", "Archive", "Lead"],
            [None, None, None, None],
            ["Bob", "B002", None, ""],
        ],
    )

    rows = ExcelPersonImportReader().read(source)

    assert [row.name for row in rows] == ["Alice", "Bob"]
    assert rows[0].identity == "A001"
    assert rows[0].department == "Archive"
    assert rows[0].note == "Lead"
    assert rows[0].row_number == 2
    assert rows[1].department is None
    assert rows[1].note is None


@requires_openpyxl
def test_excel_person_import_reader_reads_without_header(tmp_path: Path) -> None:
    """Treat the first worksheet row as data when has_header is false."""
    source = tmp_path / "people.xlsx"
    save_workbook(source, [["Alice", 1001, "Archive", None]])

    rows = ExcelPersonImportReader().read(source, has_header=False)

    assert len(rows) == 1
    assert rows[0].name == "Alice"
    assert rows[0].identity == "1001"
    assert rows[0].department == "Archive"
    assert rows[0].note is None
    assert rows[0].row_number == 1


@requires_openpyxl
def test_excel_person_import_reader_reads_named_sheet(tmp_path: Path) -> None:
    """Read rows from the worksheet selected by sheet name."""
    from openpyxl import Workbook

    source = tmp_path / "people.xlsx"
    workbook = Workbook()
    workbook.active.title = "Ignored"
    workbook.active.append(["Wrong", "W001"])
    selected = workbook.create_sheet("Students")
    selected.append(["name", "identity"])
    selected.append(["Charlie", "C003"])
    workbook.save(source)
    workbook.close()

    rows = ExcelPersonImportReader().read(source, sheet_name="Students")

    assert len(rows) == 1
    assert rows[0].name == "Charlie"
    assert rows[0].identity == "C003"
    assert rows[0].row_number == 2


@requires_openpyxl
def test_excel_person_import_reader_raises_for_missing_sheet(tmp_path: Path) -> None:
    """Surface invalid worksheet selection errors to the application layer."""
    source = tmp_path / "people.xlsx"
    save_workbook(source, [["name"], ["Alice"]])

    with pytest.raises(KeyError):
        ExcelPersonImportReader().read(source, sheet_name="Missing")