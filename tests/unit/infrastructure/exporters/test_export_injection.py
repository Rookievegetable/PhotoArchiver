"""Spreadsheet formula-injection sanitization tests (P2-003 fix).

Covers the shared ``sanitize_spreadsheet_cell`` helper plus end-to-end checks
through ``CsvExporter`` and ``ExcelExporter``: hostile user-controlled fields
must never survive as executable spreadsheet formulas, while legitimate
values (plain text, negative numbers) stay untouched.
"""

from csv import reader as csv_reader
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

import openpyxl

from photo_archiver.application.ports.exporter import ExportRow
from photo_archiver.infrastructure.exporters import CsvExporter, ExcelExporter
from photo_archiver.infrastructure.exporters._spreadsheet_safety import (
    sanitize_spreadsheet_cell,
)


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("=CMD('x')", "'=CMD('x')"),
        ("+1+1", "'+1+1"),
        ("@SUM(1)", "'@SUM(1)"),
        ("-2+3", "'-2+3"),  # non-numeric after '-' → neutralized
        ("-0.5", "-0.5"),  # legitimate negative number → kept
        ("\t=CMD", "'\t=CMD"),
        ("\r=CMD", "'\r=CMD"),
        ("Alice", "Alice"),  # plain text untouched
        ("/src/a.jpg", "/src/a.jpg"),  # paths untouched
        ("", ""),  # empty string untouched
        (0.95, 0.95),  # numbers untouched
        (None, None),
    ],
)
def test_sanitize_spreadsheet_cell_neutralizes_formula_prefixes(
    raw: str | float | None, expected: str | float | None
) -> None:
    assert sanitize_spreadsheet_cell(raw) == expected


def _row(person_name: str, note: str | None = None, confidence: float | None = 0.95) -> ExportRow:
    return ExportRow(
        person_name=person_name,
        person_department="Eng",
        person_note=note,
        photo_path="/src/a.jpg",
        photo_original_name="a.jpg",
        photo_folder="/src",
        photo_captured_at="2024-05-01",
        match_confidence=confidence,
        match_status="approved",
        archive_status="planned",
        archive_target="/archive/Alice/a.jpg",
        archive_archived_at="2024-06-01",
    )


def test_csv_exporter_neutralizes_formula_injection(tmp_path: Path) -> None:
    """A person named like a formula must not survive as one in the CSV."""
    out = tmp_path / "export.csv"
    CsvExporter().export([_row(person_name='=HYPERLINK("http://evil", "pwn")')], str(out))

    with open(out, encoding="utf-8-sig", newline="") as f:
        rows = list(csv_reader(f))

    cell = rows[1][0]  # first data row, Person Name column
    assert cell.startswith("'")
    assert not cell.startswith("=")


def test_csv_exporter_keeps_negative_numbers_and_plain_text(tmp_path: Path) -> None:
    """Sanitization must not corrupt legitimate values.

    ``-12.5`` parses as a number so it keeps its sign; dashes that are not
    leading characters are untouched. A leading ``-`` followed by non-numeric
    text (``-plain note``) *is* neutralized — that case is covered by the
    parametrized helper tests above.
    """
    out = tmp_path / "export.csv"
    CsvExporter().export([_row(person_name="Alice", note="-12.5")], str(out))

    with open(out, encoding="utf-8-sig", newline="") as f:
        rows = list(csv_reader(f))

    assert rows[1][0] == "Alice"
    assert rows[1][2] == "-12.5"


def test_excel_exporter_neutralizes_formula_injection(tmp_path: Path) -> None:
    """Excel cells must not contain raw formula-leading user data."""
    out = tmp_path / "export.xlsx"
    ExcelExporter().export([_row(person_name="=CMD|'/C calc'!A0")], str(out))

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    cell = ws.cell(row=2, column=1)  # first data row, Person Name column

    assert isinstance(cell.value, str)
    assert cell.value.startswith("'")
    assert not cell.value.startswith("=")


def test_excel_exporter_keeps_numbers_and_plain_text(tmp_path: Path) -> None:
    """Negative confidence stays numeric; plain names stay verbatim."""
    out = tmp_path / "export.xlsx"
    ExcelExporter().export([_row(person_name="Alice", confidence=-0.5)], str(out))

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=8).value == -0.5  # Match Confidence column


def test_html_exporter_template_markers_stay_data(tmp_path: Path) -> None:
    """P2-004: user data containing ``string.Template`` markers must never be
    re-substituted as template code (single-pass ``Template.substitute``)."""
    from photo_archiver.infrastructure.exporters import HtmlExporter

    out = tmp_path / "report.html"
    HtmlExporter().export([_row(person_name="$rows ${header} $cells")], str(out))
    content = out.read_text(encoding="utf-8")

    # Markers survive verbatim as escaped cell *data*…
    assert "<td>$rows ${header} $cells</td>" in content
    # …and the document structure was not rewritten by a second substitution.
    assert content.count("<table>") == 1
    assert content.count("<tbody>") == 1
