"""Tests for the extension-dispatching person import reader (P0-3).

Routing must send openpyxl-compatible workbooks to the Excel reader and
everything else to the delimited-text reader, with the port contract
(has_header / sheet_name) passed through untouched.
"""

from importlib.util import find_spec
from pathlib import Path

import pytest

from photo_archiver.infrastructure import (
    DispatchingPersonImportReader,
    ExcelPersonImportReader,
    TxtPersonImportReader,
)

requires_openpyxl = pytest.mark.skipif(find_spec("openpyxl") is None, reason="openpyxl is not installed")


def save_workbook(path: Path, rows: list[list[object | None]]) -> None:
    """Create a real Excel workbook with the provided rows."""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "People"
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def make_dispatcher() -> DispatchingPersonImportReader:
    """Build the dispatcher with the real concrete readers."""
    return DispatchingPersonImportReader(
        excel_reader=ExcelPersonImportReader(),
        text_reader=TxtPersonImportReader(),
    )


@requires_openpyxl
def test_xlsx_routes_to_excel_reader(tmp_path: Path) -> None:
    """A .xlsx workbook is parsed by the Excel reader (real openpyxl)."""
    source = tmp_path / "people.xlsx"
    save_workbook(source, [["name", "identity", "department", "note"], ["Alice", "A001", "Archive", None]])

    rows = make_dispatcher().read(source)

    assert [row.name for row in rows] == ["Alice"]
    assert rows[0].identity == "A001"


@requires_openpyxl
def test_xlsm_routes_to_excel_reader(tmp_path: Path) -> None:
    """A .xlsm workbook also routes to the Excel reader."""
    source = tmp_path / "people.xlsm"
    save_workbook(source, [["name", "identity", "department", "note"], ["Bob", "B002", None, None]])

    rows = make_dispatcher().read(source)

    assert [row.name for row in rows] == ["Bob"]


def test_txt_routes_to_text_reader(tmp_path: Path) -> None:
    """A .txt file is parsed by the delimited-text reader."""
    source = tmp_path / "people.txt"
    source.write_text(
        "name,identity,department,note\nAlice,A001,Archive,Lead\n",
        encoding="utf-8",
    )

    rows = make_dispatcher().read(source)

    assert [row.name for row in rows] == ["Alice"]
    assert rows[0].department == "Archive"


def test_csv_routes_to_text_reader(tmp_path: Path) -> None:
    """A .csv file routes to the delimited-text reader (native csv parsing)."""
    source = tmp_path / "people.csv"
    source.write_text(
        "name,identity,department,note\nBob,B002,,\n",
        encoding="utf-8",
    )

    rows = make_dispatcher().read(source)

    assert [row.name for row in rows] == ["Bob"]


@requires_openpyxl
def test_sheet_name_passthrough_reaches_excel_reader(tmp_path: Path) -> None:
    """sheet_name passes through so callers can select a named worksheet."""
    source = tmp_path / "people.xlsx"
    save_workbook(source, [["ignored-header-row"]])
    from openpyxl import load_workbook

    workbook = load_workbook(source)
    workbook.create_sheet("Roster").append(["name", "identity", "department", "note"])
    workbook["Roster"].append(["Carol", "C003", None, None])
    workbook.save(source)
    workbook.close()

    rows = make_dispatcher().read(source, sheet_name="Roster")

    assert [row.name for row in rows] == ["Carol"]


def test_has_header_false_passthrough(tmp_path: Path) -> None:
    """has_header=False passes through to the text reader (no row skipped)."""
    source = tmp_path / "people.txt"
    source.write_text("Alice,A001,Archive,\n", encoding="utf-8")

    rows = make_dispatcher().read(source, has_header=False)

    assert [row.name for row in rows] == ["Alice"]
