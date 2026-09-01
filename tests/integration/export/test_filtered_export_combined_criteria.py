"""FILTERED export linkage over combined filter criteria — Phase 9 FEAT-P9-3.

The P0 core acceptance point: the criteria the user builds in the REAL
FilterBar (status / person / date axes, any combination) rides the Phase 7
hold point (``MainWindow._current_criteria``) through the real chain

    FilterBar → criteria_changed → _on_filter_changed → _current_criteria
      → ExportDialog(active_criteria) → ExportController → QtWorkerExecutor
        → ExportTask → ExportService._gather_filtered
          → PhotoRepository.search(criteria) → list_by_photo_ids ×2
            → CsvExporter / ExcelExporter → real file

and the exported file contains exactly the criteria-matched data. Reuses the
Phase 7 boundary-double strategy (modal ExportDialog + QMessageBox.warning
recorded — Dependency/Boundary/Reason/What-remains-real documented in
``test_export_scope_integration.py``); repositories, SQLite, service, task,
controller and exporters all real.

Seeding (real SQLite, FK order, single shared Folder):

    Photo A alice_portrait.jpg  captured 2023-05-01  Alice PENDING + APPROVED, ARCHIVED
    Photo B bob_candid.jpg      captured 2024-06-15  Bob APPROVED, PLANNED
    Photo C alice_party.jpg     captured 2024-06-20  Alice REJECTED + PENDING

Per-criteria assertions follow the Phase 7 leakage-matrix discipline: the
matched photos' data must appear, every non-matched photo's data (photo row,
match rows, person row, archive row) must be absent.
"""

import pytest

pytest.importorskip("pytestqt")
pytest.importorskip("PySide6")

from csv import reader as csv_reader
from datetime import datetime
from pathlib import Path

from PySide6.QtCore import QDateTime

from photo_archiver.app import bootstrap_application
from photo_archiver.application.dtos.export import ExportScope
from photo_archiver.domain import (
    ArchiveStatus,
    Folder,
    Person,
    Photo,
    PhotoPath,
    RecognitionResult,
)
from photo_archiver.domain.entities.archive import ArchiveRecord
from photo_archiver.infrastructure.config import AppSettings
from photo_archiver.presentation.views import main_window as main_window_module
from photo_archiver.presentation.views.main_window import MainWindow

_WAIT_TERMINAL_MS = 15000


# ── Seeding (real repositories, real FK order) ──────────────────────────────


def _seed(repositories) -> dict:
    """Seed the real SQLite database; return ids the assertions need."""
    folder = Folder(path=PhotoPath("photos"), total_photos=3)
    repositories.folders.add(folder)
    photo_a = Photo(
        path=PhotoPath("photos/alice_portrait.jpg"),
        folder_id=folder.id,
        original_name="alice_portrait.jpg",
        captured_at=datetime(2023, 5, 1, 10, 0, 0),
    )
    photo_b = Photo(
        path=PhotoPath("photos/bob_candid.jpg"),
        folder_id=folder.id,
        original_name="bob_candid.jpg",
        captured_at=datetime(2024, 6, 15, 10, 0, 0),
    )
    photo_c = Photo(
        path=PhotoPath("photos/alice_party.jpg"),
        folder_id=folder.id,
        original_name="alice_party.jpg",
        captured_at=datetime(2024, 6, 20, 10, 0, 0),
    )
    repositories.photos.add(photo_a)
    repositories.photos.add(photo_b)
    repositories.photos.add(photo_c)

    alice = Person(name="Alice")
    bob = Person(name="Bob")
    repositories.people.add(alice)
    repositories.people.add(bob)

    a_pending = RecognitionResult(photo_id=photo_a.id, confidence=0.87, person_id=alice.id)  # type: ignore[arg-type]
    repositories.recognition.add(a_pending)
    a_approved = RecognitionResult(photo_id=photo_a.id, confidence=0.92, person_id=alice.id)  # type: ignore[arg-type]
    a_approved.approve()
    repositories.recognition.add(a_approved)
    b_approved = RecognitionResult(photo_id=photo_b.id, confidence=0.81, person_id=bob.id)  # type: ignore[arg-type]
    b_approved.approve()
    repositories.recognition.add(b_approved)
    c_rejected = RecognitionResult(photo_id=photo_c.id, confidence=0.44, person_id=alice.id)  # type: ignore[arg-type]
    c_rejected.reject()
    repositories.recognition.add(c_rejected)
    c_pending = RecognitionResult(photo_id=photo_c.id, confidence=0.55, person_id=alice.id)  # type: ignore[arg-type]
    repositories.recognition.add(c_pending)

    archive_root = str(repositories._connection_provider.database_path.parent)
    record_a = ArchiveRecord(
        photo_id=photo_a.id,  # type: ignore[arg-type]
        target_archive_root=archive_root,
        target_person_name="Alice",
        target_event_or_date="2023-05",
        target_original_name="alice_portrait.jpg",
        status=ArchiveStatus.PLANNED,
    )
    record_a.mark_archived()
    repositories.archive_records.add(record_a)
    repositories.archive_records.add(
        ArchiveRecord(
            photo_id=photo_b.id,  # type: ignore[arg-type]
            target_archive_root=archive_root,
            target_person_name="Bob",
            target_event_or_date="2024-06",
            target_original_name="bob_candid.jpg",
            status=ArchiveStatus.PLANNED,
        )
    )

    return {
        "photo_a_id": photo_a.id,
        "photo_b_id": photo_b.id,
        "photo_c_id": photo_c.id,
        "alice_id": alice.id,
        "bob_id": bob.id,
    }


def _make_window(qtbot, tmp_path: Path):
    """Real MainWindow over a real seeded SQLite database."""
    settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'linkage.db'}")
    settings.ensure_runtime_directories()
    context = bootstrap_application(settings)
    seeded = _seed(context.repositories)
    window = MainWindow(context)
    qtbot.addWidget(window)
    return window, context, seeded


# ── FilterBar drivers (real widgets — the user's filter construction) ───────


def _select_person(window, person_id) -> None:
    # Combo userData carries string ids (QVariant identity pitfall — see
    # FilterBar.set_persons).
    index = window._filter_bar._person_combo.findData(str(person_id))
    window._filter_bar._person_combo.setCurrentIndex(index)


def _select_status(window, index: int) -> None:
    window._filter_bar._status_combo.setCurrentIndex(index)


def _set_date(window, *, frm: tuple | None = None, to: tuple | None = None) -> None:
    if frm is not None:
        window._filter_bar._from_check.setChecked(True)
        window._filter_bar._from_edit.setDateTime(QDateTime(*frm))
    if to is not None:
        window._filter_bar._to_check.setChecked(True)
        window._filter_bar._to_edit.setDateTime(QDateTime(*to))


# ── Export boundary doubles (Phase 7 established strategy) ──────────────────


class _FakeExportDialog:
    def __init__(self, parent=None, *, output_path, scope, format_name, active_criteria=None):
        self.parent = parent
        self.active_criteria = active_criteria
        self._output_path = output_path
        self._scope = scope
        self._format_name = format_name

    def exec(self) -> int:
        return 1  # QDialog.DialogCode.Accepted

    @property
    def output_path(self):
        return self._output_path

    @property
    def scope(self):
        return self._scope

    @property
    def format_name(self):
        return self._format_name


def _stub_dialog(monkeypatch, output_path, scope, format_name) -> list:
    created: list[_FakeExportDialog] = []

    def _factory(parent=None, active_criteria=None):
        dialog = _FakeExportDialog(
            parent=parent,
            output_path=output_path,
            scope=scope,
            format_name=format_name,
            active_criteria=active_criteria,
        )
        created.append(dialog)
        return dialog

    monkeypatch.setattr(main_window_module, "ExportDialog", _factory)
    return created


def _run_export(qtbot, window, output_path, format_name, monkeypatch):
    created = _stub_dialog(monkeypatch, output_path, ExportScope.FILTERED, format_name)
    window._export_action.trigger()
    qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)
    assert window._status_label.text() == "export complete"
    assert output_path.exists()
    assert created[0].active_criteria is window._current_criteria  # F5 forwarding
    return created[0].active_criteria


def _read_csv(output_path: Path) -> list[list[str]]:
    with open(output_path, encoding="utf-8-sig", newline="") as handle:
        return list(csv_reader(handle))[1:]


def _sections(rows: list[list[str]]) -> dict:
    sections: dict = {"people": [], "photos": [], "matches": [], "archive": []}
    for row in rows:
        _, _, _, photo_path, _, _, _, _, match_status, archive_status, _, _ = row
        if match_status:
            sections["matches"].append(row)
        elif archive_status:
            sections["archive"].append(row)
        elif ".jpg" in photo_path:
            sections["photos"].append(row)
        else:
            sections["people"].append(row)
    return sections


# ── Linkage tests ────────────────────────────────────────────────────────────


def test_status_only_filtered_export_csv(qtbot, tmp_path, monkeypatch) -> None:
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_status(window, 1)  # Pending
    assert window._current_criteria is not None
    assert window._current_criteria.match_status.value == "pending"

    criteria = _run_export(qtbot, window, tmp_path / "status.csv", "csv", monkeypatch)
    assert criteria.match_status.value == "pending"

    sections = _sections(_read_csv(tmp_path / "status.csv"))
    # Main set: A + C (both carry a PENDING recognition); B excluded.
    assert {row[4] for row in sections["photos"]} == {"alice_portrait.jpg", "alice_party.jpg"}
    # matches = ALL statuses of the main set (contract §3/F4).
    assert {row[8] for row in sections["matches"]} == {"pending", "approved", "rejected"}
    assert str(seeded["photo_b_id"]) not in {row[3] for row in sections["matches"]}
    # people derived from matches: only Alice (B's Bob must not leak).
    assert {row[0] for row in sections["people"]} == {"Alice"}
    # archive: only A's archived record; B's PLANNED must not leak.
    assert len(sections["archive"]) == 1
    assert sections["archive"][0][9] == "archived"
    flat = {cell for row in _read_csv(tmp_path / "status.csv") for cell in row}
    assert "bob_candid.jpg" not in flat and "Bob" not in flat and "planned" not in flat


def test_person_only_filtered_export_csv(qtbot, tmp_path, monkeypatch) -> None:
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_person(window, seeded["bob_id"])
    assert window._current_criteria.person_id == seeded["bob_id"]

    _run_export(qtbot, window, tmp_path / "person.csv", "csv", monkeypatch)

    sections = _sections(_read_csv(tmp_path / "person.csv"))
    assert {row[4] for row in sections["photos"]} == {"bob_candid.jpg"}
    assert {row[8] for row in sections["matches"]} == {"approved"}
    assert {row[0] for row in sections["people"]} == {"Bob"}
    assert len(sections["archive"]) == 1
    assert sections["archive"][0][9] == "planned"
    flat = {cell for row in _read_csv(tmp_path / "person.csv") for cell in row}
    assert "alice_portrait.jpg" not in flat and "alice_party.jpg" not in flat
    assert "Alice" not in flat  # wrong person must not leak anywhere


def test_date_only_filtered_export_xlsx(qtbot, tmp_path, monkeypatch) -> None:
    from openpyxl import load_workbook

    window, _, seeded = _make_window(qtbot, tmp_path)
    _set_date(window, frm=(2024, 1, 1, 0, 0, 0), to=(2024, 12, 31, 23, 59, 59))
    assert window._current_criteria.captured_from is not None

    _run_export(qtbot, window, tmp_path / "date.xlsx", "xlsx", monkeypatch)

    workbook = load_workbook(str(tmp_path / "date.xlsx"))
    sheet = workbook.active
    assert sheet is not None and sheet.title == "Export"
    values = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    sections = {"people": [], "photos": [], "matches": [], "archive": []}
    for row in values:
        _, _, _, photo_path, _, _, _, _, match_status, archive_status, _, _ = [
            "" if cell is None else str(cell) for cell in row
        ]
        if match_status:
            sections["matches"].append(row)
        elif archive_status:
            sections["archive"].append(row)
        elif ".jpg" in photo_path:
            sections["photos"].append(row)
        else:
            sections["people"].append(row)
    # Main set: B + C (both captured 2024); A (2023) excluded.
    assert {str(row[4]) for row in sections["photos"]} == {"bob_candid.jpg", "alice_party.jpg"}
    assert {str(row[8]) for row in sections["matches"]} == {"approved", "rejected", "pending"}
    assert {str(row[0]) for row in sections["people"]} == {"Bob", "Alice"}
    flat = {str(cell) for row in values for cell in row}
    assert "alice_portrait.jpg" not in flat  # wrong date range must not leak
    assert str(seeded["photo_a_id"]) not in flat


def test_person_and_date_filtered_export_csv(qtbot, tmp_path, monkeypatch) -> None:
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_person(window, seeded["alice_id"])
    _set_date(window, frm=(2024, 6, 1, 0, 0, 0))  # from-only gate
    criteria = window._current_criteria
    assert criteria.person_id == seeded["alice_id"] and criteria.captured_from is not None

    _run_export(qtbot, window, tmp_path / "person_date.csv", "csv", monkeypatch)

    sections = _sections(_read_csv(tmp_path / "person_date.csv"))
    # Main set: only C (Alice's 2024 photo). A excluded by date, B by person.
    assert {row[4] for row in sections["photos"]} == {"alice_party.jpg"}
    assert {row[8] for row in sections["matches"]} == {"rejected", "pending"}
    assert str(seeded["photo_a_id"]) not in {row[3] for row in sections["matches"]}
    assert str(seeded["photo_b_id"]) not in {row[3] for row in sections["matches"]}
    flat = {cell for row in _read_csv(tmp_path / "person_date.csv") for cell in row}
    assert "alice_portrait.jpg" not in flat and "bob_candid.jpg" not in flat
    assert "0.87" not in flat and "0.92" not in flat and "0.81" not in flat


def test_person_date_status_triple_filtered_export_csv(qtbot, tmp_path, monkeypatch) -> None:
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_person(window, seeded["alice_id"])
    _set_date(window, frm=(2024, 1, 1, 0, 0, 0), to=(2024, 12, 31, 23, 59, 59))
    _select_status(window, 3)  # Rejected
    criteria = window._current_criteria
    assert criteria.person_id == seeded["alice_id"]
    assert criteria.match_status.value == "rejected"
    assert criteria.captured_to is not None

    _run_export(qtbot, window, tmp_path / "triple.csv", "csv", monkeypatch)

    sections = _sections(_read_csv(tmp_path / "triple.csv"))
    assert {row[4] for row in sections["photos"]} == {"alice_party.jpg"}
    # matches stay ALL statuses of the main set (F4): rejected + pending of C.
    assert {row[8] for row in sections["matches"]} == {"rejected", "pending"}
    flat = {cell for row in _read_csv(tmp_path / "triple.csv") for cell in row}
    assert "alice_portrait.jpg" not in flat and "bob_candid.jpg" not in flat


def test_no_matches_filtered_export_is_header_only(qtbot, tmp_path, monkeypatch) -> None:
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_person(window, seeded["bob_id"])
    _set_date(window, frm=(2023, 1, 1, 0, 0, 0), to=(2023, 12, 31, 23, 59, 59))
    assert window._current_criteria is not None  # Bob has no 2023 photo

    _run_export(qtbot, window, tmp_path / "empty.csv", "csv", monkeypatch)

    rows = _read_csv(tmp_path / "empty.csv")
    assert rows == []  # header only — honest empty main set


def test_export_always_uses_latest_criteria_not_stale(qtbot, tmp_path, monkeypatch) -> None:
    """Two consecutive exports: the second must reflect the NEW filter only."""
    window, _, seeded = _make_window(qtbot, tmp_path)
    _select_person(window, seeded["bob_id"])
    first = _run_export(qtbot, window, tmp_path / "first.csv", "csv", monkeypatch)
    assert first.person_id == seeded["bob_id"]

    _select_person(window, seeded["alice_id"])
    _set_date(window, frm=(2024, 6, 1, 0, 0, 0))
    second = _run_export(qtbot, window, tmp_path / "second.csv", "csv", monkeypatch)
    assert second.person_id == seeded["alice_id"]
    assert second.captured_from is not None

    first_sections = _sections(_read_csv(tmp_path / "first.csv"))
    second_sections = _sections(_read_csv(tmp_path / "second.csv"))
    assert {row[4] for row in first_sections["photos"]} == {"bob_candid.jpg"}
    assert {row[4] for row in second_sections["photos"]} == {"alice_party.jpg"}
    # criteria leakage across runs would merge the two main sets — forbid it.
    flat_first = {cell for row in _read_csv(tmp_path / "first.csv") for cell in row}
    assert "alice_party.jpg" not in flat_first
