"""Export UI-to-file closed-loop integration (Phase 5 Commit 2, FEATURE-002).

Runs the REAL production chain from the Commit 1 UI entry point:

    MainWindow "Export Data" QAction
      → _on_export_clicked (real handler)
        → ExportController.export (real controller, real exporters dict)
          → QtWorkerExecutor / QThreadPool (real worker, queued signals)
            → ExportTask.execute (real task, two-phase progress)
              → ExportService.export (real service, real repositories)
                → CsvExporter / ExcelExporter (real infrastructure exporters)
                  → real SQLite reads (alembic schema, foreign keys ON)
                    → real output file on the filesystem

Nothing in the chain is mocked: Controller, Task, Service, Exporter,
SQLite repositories, and the output filesystem are all production objects.
The ONLY replaced boundaries are two non-deterministic modal UI surfaces:

    Dependency:           ExportDialog (module attr ``main_window.ExportDialog``)
    Boundary:             modal user interaction (headless test cannot drive a
                          real modal exec loop deterministically)
    Why replaced:         real modal exec would block the Qt event loop;
                          the real dialog's property contract (scope /
                          output_path / format_name after accept) is already
                          covered by the Commit 1 unit suite
    What remains real:    the handler's dialog-gated flow, and everything
                          after ``dialog.exec()`` returns Accepted

    Dependency:           QMessageBox.warning (recorded double)
    Boundary:             modal error popup on the failure path
    Why replaced:         a real modal would block the test event loop
    What remains real:    the full failure chain TaskFailed → shared error
                          surface → action re-enable

Data design (business-level included/excluded semantics, since
``ExportService._gather_data`` treats every scope as ALL — CURRENT_BATCH and
FILTERED are documented stubs, FEATURE-004):

    included: 2 people, 2 photos, 1 APPROVED match, 1 ARCHIVED record = 6 rows
    excluded: the PENDING match (Bob↔bob_candid.jpg) and the REJECTED match
              (bob_candid.jpg↔Alice) never enter the matches section, because
              only ``list_approved_by_person`` rows are gathered.
"""

import pytest

pytest.importorskip("pytestqt")
pytest.importorskip("PySide6")

from csv import reader as csv_reader
from pathlib import Path

from photo_archiver.app import bootstrap_application  # noqa: F401
from photo_archiver.application.dtos.export import ExportScope
from photo_archiver.domain import (
    ArchiveStatus,
    Folder,
    Person,
    Photo,
    PhotoPath,
    RecognitionResult,
)
from photo_archiver.domain.entities.archive import ArchiveRecord
from photo_archiver.infrastructure.config import AppSettings
from photo_archiver.presentation.views import main_window as main_window_module
from photo_archiver.presentation.views.main_window import MainWindow
from photo_archiver.workers import QtWorkerExecutor

_WAIT_TERMINAL_MS = 15000


def _seed_sqlite(repositories) -> dict:
    """Seed the real SQLite database through the real repositories (FK order).

    Returns the ids the CSV assertions need (photo ids, folder id). Insert
    order honours the production foreign keys: folders → photos → people →
    recognition_results → archive_records.
    """
    folder = Folder(path=PhotoPath("photos"), total_photos=2)
    repositories.folders.add(folder)

    photo_a = Photo(path=PhotoPath("photos/alice_portrait.jpg"), folder_id=folder.id, original_name="alice_portrait.jpg")
    photo_b = Photo(path=PhotoPath("photos/bob_candid.jpg"), folder_id=folder.id, original_name="bob_candid.jpg")
    repositories.photos.add(photo_a)
    repositories.photos.add(photo_b)

    alice = Person(name="Alice", department="Engineering")
    bob = Person(name="Bob", note="vip")
    repositories.people.add(alice)
    repositories.people.add(bob)

    # R1: APPROVED — the only match row the export may contain.
    approved = RecognitionResult(photo_id=photo_a.id, confidence=0.92, person_id=alice.id)  # type: ignore[arg-type]
    approved.approve()
    repositories.recognition.add(approved)
    # R2: PENDING — must be excluded from the matches section.
    pending = RecognitionResult(photo_id=photo_b.id, confidence=0.87, person_id=bob.id)  # type: ignore[arg-type]
    repositories.recognition.add(pending)
    # R3: REJECTED — must be excluded from the matches section.
    rejected = RecognitionResult(photo_id=photo_b.id, confidence=0.55, person_id=alice.id)  # type: ignore[arg-type]
    rejected.reject()
    repositories.recognition.add(rejected)

    record = ArchiveRecord(
        photo_id=photo_a.id,  # type: ignore[arg-type]
        target_archive_root=str(repositories._connection_provider.database_path.parent),
        target_person_name="Alice",
        target_event_or_date="2024-01",
        target_original_name="alice_portrait.jpg",
        status=ArchiveStatus.PLANNED,
    )
    record.mark_archived()  # finalize BEFORE add so the row persists ARCHIVED + timestamp
    repositories.archive_records.add(record)

    return {
        "photo_a_id": photo_a.id,
        "photo_b_id": photo_b.id,
        "folder_id": folder.id,
    }


def _make_window(qtbot, tmp_path: Path, *, seed: bool = True) -> MainWindow:
    """Build a real MainWindow over a real tmp SQLite context (Commit 1 pattern).

    ``seed=True`` (default) populates the database through the real
    repositories; ``seed=False`` leaves the database empty for empty-database
    scenarios.
    """
    settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'export_ui.db'}")
    settings.ensure_runtime_directories()
    context = bootstrap_application(settings)
    seeded = _seed_sqlite(context.repositories) if seed else {}
    window = MainWindow(context)
    qtbot.addWidget(window)
    return window, context, seeded


class _FakeExportDialog:
    """Modal-dialog boundary double — mirrors the real dialog's read surface.

    The real ``ExportDialog`` (scope / output_path / format_name properties,
    ``exec()`` accept gate) is unit-covered in Commit 1; here only the
    modal boundary is replaced so the production handler flow stays real.
    """

    def __init__(self, parent=None, *, output_path: Path, scope: ExportScope, format_name: str) -> None:
        self.parent = parent
        self._output_path = output_path
        self._scope = scope
        self._format_name = format_name

    def exec(self) -> int:
        """Return Accepted immediately (the user confirmed the export)."""
        return 1  # QDialog.DialogCode.Accepted

    @property
    def output_path(self) -> Path:
        """Return the validated output path the fake user chose."""
        return self._output_path

    @property
    def scope(self) -> ExportScope:
        """Return the selected export scope."""
        return self._scope

    @property
    def format_name(self) -> str:
        """Return the canonical format name for the exporter lookup."""
        return self._format_name


def _stub_dialog(monkeypatch, output_path: Path, scope: ExportScope, format_name: str) -> list:
    """Patch the modal boundary in the window module; capture created dialogs."""
    created: list[_FakeExportDialog] = []

    def _factory(parent=None) -> _FakeExportDialog:
        dialog = _FakeExportDialog(
            parent=parent, output_path=output_path, scope=scope, format_name=format_name
        )
        created.append(dialog)
        return dialog

    monkeypatch.setattr(main_window_module, "ExportDialog", _factory)
    return created


def _rows_by_section(rows: list[list[str]]) -> dict:
    """Classify flattened CSV rows by section (order-stable assertions).

    The repositories order same-second inserts by uuid id, so row order is
    not deterministic across runs. Classification keys on the section
    markers the real ``_flatten`` produces:
      person row:  empty photo_path/match_status/archive_status
      photo row:   .jpg photo_path, empty match/archive status
      match row:   non-empty match_status
      archive row: non-empty archive_status
    """
    sections: dict = {"people": [], "photos": [], "matches": [], "archive": []}
    for row in rows:
        _, _, _, photo_path, _, _, _, _, match_status, archive_status, _, _ = row
        if match_status:
            sections["matches"].append(row)
        elif archive_status:
            sections["archive"].append(row)
        elif ".jpg" in photo_path:
            sections["photos"].append(row)
        else:
            sections["people"].append(row)
    return sections


def test_csv_ui_trigger_exports_real_sqlite_rows_to_file(qtbot, tmp_path, monkeypatch) -> None:
    """Full closed loop: UI click → real chain → real SQLite → real CSV file.

    Proves the export wrote production-gathered SQLite rows through the real
    CsvExporter onto the filesystem, with PENDING/REJECTED matches excluded.
    """
    window, context, seeded = _make_window(qtbot, tmp_path)
    # Assembly evidence: the window drives the production controller/executor.
    assert window._export_controller is context.export_controller
    assert window._export_controller._executor is context.worker_executor
    assert isinstance(context.worker_executor, QtWorkerExecutor)

    output_path = tmp_path / "export.csv"
    created = _stub_dialog(monkeypatch, output_path, ExportScope.ALL, "csv")

    window._export_action.trigger()

    assert created and created[0].parent is window
    # Terminal observation: completed/failed handlers re-enable the action.
    qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)
    assert window._status_label.text() == "export complete"

    # ── Real file ──────────────────────────────────────────────────────────
    assert output_path.exists()
    with open(output_path, encoding="utf-8-sig", newline="") as handle:
        parsed = list(csv_reader(handle))

    expected_headers = [
        "person_name", "department", "note", "photo_path", "original_name",
        "folder", "captured_at", "match_confidence", "match_status",
        "archive_status", "archive_target", "archived_at",
    ]
    assert parsed[0] == expected_headers
    data_rows = parsed[1:]
    assert len(data_rows) == 6  # 2 people + 2 photos + 1 approved match + 1 archive

    sections = _rows_by_section(data_rows)
    # ── People section (fields read from the real SQLite people table) ─────
    alice_rows = [r for r in sections["people"] if r[0] == "Alice"]
    bob_rows = [r for r in sections["people"] if r[0] == "Bob"]
    assert len(alice_rows) == 1 and len(bob_rows) == 1
    assert alice_rows[0][1] == "Engineering"  # department column
    assert bob_rows[0][2] == "vip"  # note column

    # ── Photos section (fields read from the real SQLite photos table) ─────
    by_original = {r[4]: r for r in sections["photos"]}
    assert set(by_original) == {"alice_portrait.jpg", "bob_candid.jpg"}
    row_a = by_original["alice_portrait.jpg"]
    assert "alice_portrait.jpg" in row_a[3]  # str(PhotoPath) carries the file name
    assert row_a[5] == str(seeded["folder_id"])  # folder column = real folder uuid

    # ── Matches section (only the APPROVED result from list_approved_by_person)
    assert len(sections["matches"]) == 1
    match_row = sections["matches"][0]
    assert match_row[7] == "0.92"  # match_confidence
    assert match_row[8] == "approved"  # match_status
    assert match_row[3] == str(seeded["photo_a_id"])  # flatten maps photo_id here

    # ── Excluded records: PENDING/REJECTED never exported ───────────────────
    flat_cells = {cell for row in data_rows for cell in row}
    assert "pending" not in flat_cells
    assert "rejected" not in flat_cells

    # ── Archive section (row read from the real archive_records table) ──────
    assert len(sections["archive"]) == 1
    archive_row = sections["archive"][0]
    assert archive_row[0] == "Alice"  # target_person_name
    assert archive_row[9] == "archived"
    assert archive_row[10].endswith("alice_portrait.jpg")
    assert archive_row[11] != ""  # archived_at timestamp persisted


def test_xlsx_ui_trigger_produces_openable_workbook_from_sqlite_rows(qtbot, tmp_path, monkeypatch) -> None:
    """Same real chain with the xlsx format: the workbook opens and holds the rows."""
    from openpyxl import load_workbook

    window, _, seeded = _make_window(qtbot, tmp_path)
    output_path = tmp_path / "export.xlsx"
    _stub_dialog(monkeypatch, output_path, ExportScope.ALL, "xlsx")

    window._export_action.trigger()
    qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)

    assert output_path.exists()
    workbook = load_workbook(str(output_path))
    sheet = workbook.active
    assert sheet is not None and sheet.title == "Export"
    assert sheet.max_row == 7  # header + 6 data rows
    assert sheet.cell(row=1, column=1).value == "Person Name"

    values = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    people_names = {
        row[0]
        for row in values
        if row[3] in ("", None) and row[8] in ("", None) and row[9] in ("", None)
    }
    assert {"Alice", "Bob"} <= people_names
    assert any(row[8] == "approved" and float(row[7]) == 0.92 for row in values)
    assert any(
        row[9] == "archived" and str(row[10]).endswith("alice_portrait.jpg") and row[11] is not None
        for row in values
    )
    flat_cells = {str(cell) for row in values for cell in row}
    assert "pending" not in flat_cells and "rejected" not in flat_cells
    assert str(seeded["photo_a_id"]) in {str(row[3]) for row in values}  # match row maps photo_id


def test_failed_export_surfaces_taskfailed_through_ui(qtbot, tmp_path, monkeypatch) -> None:
    """Failure path: unwritable output → TaskFailed → real UI failure surface."""
    window, _, _ = _make_window(qtbot, tmp_path)
    from PySide6.QtWidgets import QMessageBox

    dialogs: list[tuple[str, str]] = []

    def _warning(parent, title: str, message: str) -> int:
        dialogs.append((title, message))
        return 0

    monkeypatch.setattr(QMessageBox, "warning", staticmethod(_warning))
    # Writing to the directory itself is an OSError on every platform.
    _stub_dialog(monkeypatch, tmp_path, ExportScope.ALL, "csv")

    window._export_action.trigger()
    qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)

    assert window._status_label.text() == "export failed."
    assert window._progress.value() == 0
    assert dialogs, "QMessageBox.warning must surface the TaskFailed event"
    assert dialogs[0][0] == "Export Failed"
    assert dialogs[0][1] != ""


def test_empty_database_export_produces_header_only_file(qtbot, tmp_path, monkeypatch) -> None:
    """Empty-database closed loop (baseline Commit 2 plan): safe UI export.

    An empty real SQLite database must still produce a valid header-only CSV
    through the full UI chain — no crash, no spurious rows, terminal state
    reached, action re-enabled.
    """
    window, _, _ = _make_window(qtbot, tmp_path, seed=False)  # empty database
    output_path = tmp_path / "empty.csv"
    _stub_dialog(monkeypatch, output_path, ExportScope.ALL, "csv")

    window._export_action.trigger()
    qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)

    assert window._status_label.text() == "export complete"
    assert output_path.exists()
    with open(output_path, encoding="utf-8-sig", newline="") as handle:
        parsed = list(csv_reader(handle))
    assert len(parsed) == 1  # header only, zero data rows
    assert parsed[0][0] == "person_name" and parsed[0][11] == "archived_at"



