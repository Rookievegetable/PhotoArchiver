"""Export scope integration closure (Phase 7 Commit 4, FEATURE-004).

Real-chain verification of the three export scopes over a REAL SQLite
database (alembic schema, foreign keys ON), per the authorized contract
``docs/health-check/PHASE_7_SCOPE_CONTRACT_REVISION.md``:

    MainWindow "Export Data" QAction (UI tests)
      → _on_export_clicked → ExportController.export (real)
        → QtWorkerExecutor / QThreadPool (real, queued signals)
          → ExportTask.execute (real, also exercised directly headless)
            → ExportService._gather_data scope dispatch (real)
              → PhotoRepository.search(criteria)           (real SQL, FILTERED main set)
              → RecognitionRepository.list_by_photo_ids    (real SQL)
              → ArchiveRecordRepository.list_by_photo_ids  (real SQL)
                → CsvExporter / ExcelExporter (real infrastructure)
                  → real output file on the filesystem, re-parsed with
                    csv.reader / openpyxl and asserted item by item

Nothing in the business chain is mocked: repositories, SQLite, services,
tasks, controller, exporters and the output filesystem are production
objects. The ONLY replaced boundaries are the same two non-deterministic
modal UI surfaces already established by the Phase 5/Commit 1–3 suites:

    Dependency:           ExportDialog (module attr ``main_window.ExportDialog``)
    Boundary:             modal user interaction (headless test cannot drive a
                          real modal exec loop deterministically)
    Why replaced:         a real modal exec would block the Qt event loop;
                          the real dialog's property contract and the F3
                          enable/disable logic are unit-covered in Commit 3
    What remains real:    the handler's dialog-gated flow, the F5 criteria
                          forwarding (the fake records ``active_criteria``),
                          and everything after ``dialog.exec()`` accepts

    Dependency:           QMessageBox.warning (recorded double)
    Boundary:             modal error popup on the failure path
    Why replaced:         a real modal would block the test event loop
    What remains real:    the full failure chain TaskFailed → shared error
                          surface → action re-enable

Data design (leakage matrix — every section proves both inclusion AND
exclusion; FK-true seeding order folders → photos → people →
recognition_results → archive_records, one shared Folder record so the
UNIQUE(raw_path, path_base) constraint is exercised the way production
uses it):

    Photo A "alice_portrait.jpg"  — MATCHES the filter (Status=Pending):
        1 PENDING  recognition (Alice, 0.87)  → must appear in FILTERED matches
        1 APPROVED recognition (Alice, 0.92)  → must appear in FILTERED matches
        1 ARCHIVED record                        → must appear in FILTERED archive
    Photo B "bob_candid.jpg"      — does NOT match (only APPROVED/REJECTED):
        1 APPROVED recognition (Bob, 0.81)    → must NOT leak anywhere
        1 PLANNED archive record                 → must NOT leak anywhere

    FILTERED (match_status=PENDING) therefore exports: 1 photo (A),
    2 matches (A's pending + approved — ALL statuses by contract §3/F4),
    1 person (Alice, derived from matches), 1 archive record (A's) = 5 rows.
    ALL on the same database exports the historical approved-only catalog
    (2 people + 2 photos + 2 approved matches + 2 archive records = 8 rows,
    pending excluded) — the intentional ALL/FILTERED asymmetry of §3/F4.
"""

import pytest

pytest.importorskip("pytestqt")
pytest.importorskip("PySide6")

from csv import reader as csv_reader
from pathlib import Path

from photo_archiver.app import bootstrap_application
from photo_archiver.application.dtos.export import ExportScope
from photo_archiver.domain import (
    ArchiveStatus,
    Folder,
    MatchStatus,
    Person,
    Photo,
    PhotoPath,
    PhotoSearchCriteria,
    RecognitionResult,
)
from photo_archiver.domain.entities.archive import ArchiveRecord
from photo_archiver.infrastructure.config import AppSettings
from photo_archiver.infrastructure.exporters import CsvExporter, ExcelExporter
from photo_archiver.presentation.views import main_window as main_window_module
from photo_archiver.presentation.views.main_window import MainWindow
from photo_archiver.workers import ExportTask

_WAIT_TERMINAL_MS = 15000

_PENDING_CRITERIA = PhotoSearchCriteria(match_status=MatchStatus.PENDING)


# ── Seeding (real repositories, real FK order, real UNIQUE constraint) ──────


def _seed_scope_database(repositories) -> dict:
    """Seed the real SQLite database through the real repositories.

    Returns the ids the file assertions need. The same folder path backs both
    photos — a second ``add`` for the same path would violate the production
    UNIQUE(raw_path, path_base) constraint, so the Folder record is created
    exactly once and reused.
    """
    folder = Folder(path=PhotoPath("photos"), total_photos=2)
    repositories.folders.add(folder)

    photo_a = Photo(
        path=PhotoPath("photos/alice_portrait.jpg"),
        folder_id=folder.id,
        original_name="alice_portrait.jpg",
    )
    photo_b = Photo(
        path=PhotoPath("photos/bob_candid.jpg"),
        folder_id=folder.id,
        original_name="bob_candid.jpg",
    )
    repositories.photos.add(photo_a)
    repositories.photos.add(photo_b)

    alice = Person(name="Alice", department="Engineering")
    bob = Person(name="Bob", note="vip")
    repositories.people.add(alice)
    repositories.people.add(bob)

    # Photo A: one PENDING + one APPROVED recognition (both Alice).
    pending_a = RecognitionResult(photo_id=photo_a.id, confidence=0.87, person_id=alice.id)  # type: ignore[arg-type]
    repositories.recognition.add(pending_a)
    approved_a = RecognitionResult(photo_id=photo_a.id, confidence=0.92, person_id=alice.id)  # type: ignore[arg-type]
    approved_a.approve()
    repositories.recognition.add(approved_a)
    # Photo B: one APPROVED (Bob) + one REJECTED — no PENDING, so the
    # match_status=PENDING filter must keep B out of the main set.
    approved_b = RecognitionResult(photo_id=photo_b.id, confidence=0.81, person_id=bob.id)  # type: ignore[arg-type]
    approved_b.approve()
    repositories.recognition.add(approved_b)
    rejected_b = RecognitionResult(photo_id=photo_b.id, confidence=0.44, person_id=alice.id)  # type: ignore[arg-type]
    rejected_b.reject()
    repositories.recognition.add(rejected_b)

    archive_root = str(repositories._connection_provider.database_path.parent)
    record_a = ArchiveRecord(
        photo_id=photo_a.id,  # type: ignore[arg-type]
        target_archive_root=archive_root,
        target_person_name="Alice",
        target_event_or_date="2024-01",
        target_original_name="alice_portrait.jpg",
        status=ArchiveStatus.PLANNED,
    )
    record_a.mark_archived()  # finalize BEFORE add so ARCHIVED + timestamp persist
    repositories.archive_records.add(record_a)
    record_b = ArchiveRecord(
        photo_id=photo_b.id,  # type: ignore[arg-type]
        target_archive_root=archive_root,
        target_person_name="Bob",
        target_event_or_date="2024-02",
        target_original_name="bob_candid.jpg",
        status=ArchiveStatus.PLANNED,
    )
    repositories.archive_records.add(record_b)

    return {
        "folder_id": folder.id,
        "photo_a_id": photo_a.id,
        "photo_b_id": photo_b.id,
    }


def _make_context(tmp_path: Path):
    """Bootstrap a real application context over a real tmp SQLite database."""
    settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'scope_export.db'}")
    settings.ensure_runtime_directories()
    context = bootstrap_application(settings)
    seeded = _seed_scope_database(context.repositories)
    return context, seeded


# ── UI boundary doubles (Commit 1–3 established strategy) ────────────────────


class _FakeExportDialog:
    """Modal-dialog boundary double — mirrors the real dialog's read surface.

    Also records the ``active_criteria`` the production handler forwarded, so
    the F5 criteria-passing contract is asserted against the real chain.
    """

    def __init__(
        self,
        parent=None,
        *,
        output_path: Path,
        scope: ExportScope,
        format_name: str,
        active_criteria: PhotoSearchCriteria | None = None,
    ) -> None:
        self.parent = parent
        self.active_criteria = active_criteria
        self._output_path = output_path
        self._scope = scope
        self._format_name = format_name

    def exec(self) -> int:
        return 1  # QDialog.DialogCode.Accepted

    @property
    def output_path(self) -> Path:
        return self._output_path

    @property
    def scope(self) -> ExportScope:
        return self._scope

    @property
    def format_name(self) -> str:
        return self._format_name


def _stub_dialog(monkeypatch, output_path: Path, scope: ExportScope, format_name: str) -> list:
    """Patch the modal boundary in the window module; capture created dialogs."""
    created: list[_FakeExportDialog] = []

    def _factory(parent=None, active_criteria=None) -> _FakeExportDialog:
        dialog = _FakeExportDialog(
            parent=parent,
            output_path=output_path,
            scope=scope,
            format_name=format_name,
            active_criteria=active_criteria,
        )
        created.append(dialog)
        return dialog

    monkeypatch.setattr(main_window_module, "ExportDialog", _factory)
    return created


# ── CSV parsing helpers ──────────────────────────────────────────────────────


def _read_csv_rows(output_path: Path) -> tuple[list[str], list[list[str]]]:
    """Re-parse a written CSV with csv.reader; return (header, data rows)."""
    with open(output_path, encoding="utf-8-sig", newline="") as handle:
        parsed = list(csv_reader(handle))
    return parsed[0], parsed[1:]


def _rows_by_section(rows: list[list[str]]) -> dict:
    """Classify flattened CSV rows by section (order-stable, see closed loop).

    person row:  empty photo_path/match_status/archive_status
    photo row:   .jpg photo_path, empty match/archive status
    match row:   non-empty match_status (photo_path column carries photo_id)
    archive row: non-empty archive_status
    """
    sections: dict = {"people": [], "photos": [], "matches": [], "archive": []}
    for row in rows:
        _, _, _, photo_path, _, _, _, _, match_status, archive_status, _, _ = row
        if match_status:
            sections["matches"].append(row)
        elif archive_status:
            sections["archive"].append(row)
        elif ".jpg" in photo_path:
            sections["photos"].append(row)
        else:
            sections["people"].append(row)
    return sections


_EXPECTED_HEADERS = [
    "person_name", "department", "note", "photo_path", "original_name",
    "folder", "captured_at", "match_confidence", "match_status",
    "archive_status", "archive_target", "archived_at",
]


# ── A. FILTERED CSV — full real UI → SQLite → CSV chain ─────────────────────


class TestFilteredCsv:
    """FILTERED scope through the real MainWindow → controller → task chain."""

    def test_filtered_csv_ui_chain_exports_only_criteria_matched_data(
        self, qtbot, tmp_path, monkeypatch,
    ) -> None:
        """FILTERED + held criteria exports exactly the criteria main set.

        The criteria flows: FilterBar handler → ``_current_criteria`` hold
        point → dialog ``active_criteria`` → controller → task → service →
        ``PhotoRepository.search`` (real SQL) → derived sections → real CSV.
        """
        settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'filtered_csv.db'}")
        settings.ensure_runtime_directories()
        context = bootstrap_application(settings)
        seeded = _seed_scope_database(context.repositories)
        window = MainWindow(context)
        qtbot.addWidget(window)

        # User sets a filter first: the real handler updates the hold point.
        window._on_filter_changed(_PENDING_CRITERIA)
        assert window._current_criteria is _PENDING_CRITERIA

        output_path = tmp_path / "filtered.csv"
        created = _stub_dialog(monkeypatch, output_path, ExportScope.FILTERED, "csv")

        window._export_action.trigger()

        # F5: the handler forwarded the held snapshot into the dialog.
        assert created and created[0].active_criteria is _PENDING_CRITERIA
        qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)
        assert window._status_label.text() == "export complete"

        # ── Real file, re-parsed with csv.reader ───────────────────────────
        assert output_path.exists()
        headers, data_rows = _read_csv_rows(output_path)
        assert headers == _EXPECTED_HEADERS

        sections = _rows_by_section(data_rows)

        # photos = criteria main set: ONLY Photo A (search matched pending).
        assert len(sections["photos"]) == 1
        photo_row = sections["photos"][0]
        assert "alice_portrait.jpg" in photo_row[3]
        assert photo_row[4] == "alice_portrait.jpg"
        assert photo_row[5] == str(seeded["folder_id"])

        # matches = ALL statuses of the main set (§3/F4): Photo A's PENDING
        # (0.87) AND APPROVED (0.92), both bound to Photo A's id.
        assert len(sections["matches"]) == 2
        match_statuses = {row[8]: row for row in sections["matches"]}
        assert set(match_statuses) == {"pending", "approved"}
        assert match_statuses["pending"][7] == "0.87"
        assert match_statuses["approved"][7] == "0.92"
        assert {row[3] for row in sections["matches"]} == {str(seeded["photo_a_id"])}

        # people = derived from the matches: Alice only, with real columns.
        assert [row[0] for row in sections["people"]] == ["Alice"]
        assert sections["people"][0][1] == "Engineering"

        # archive_records = main set history: Photo A's ARCHIVED record only.
        assert len(sections["archive"]) == 1
        archive_row = sections["archive"][0]
        assert archive_row[0] == "Alice"
        assert archive_row[9] == "archived"
        assert archive_row[10].endswith("alice_portrait.jpg")
        assert archive_row[11] != ""

        # ── Leakage: nothing from Photo B may appear anywhere ──────────────
        flat_cells = {cell for row in data_rows for cell in row}
        assert str(seeded["photo_b_id"]) not in flat_cells  # B's match rows
        assert "bob_candid.jpg" not in flat_cells  # B's photo + archive rows
        assert "Bob" not in flat_cells  # B's person row + archive target
        assert "0.81" not in flat_cells  # B's approved recognition
        assert "0.44" not in flat_cells  # B's rejected recognition
        assert "planned" not in flat_cells  # B's PLANNED archive status
        assert "rejected" not in flat_cells  # B's REJECTED recognition

        # Summary says 5 rows: 1 person + 1 photo + 2 matches + 1 archive.
        assert len(data_rows) == 5


# ── B. FILTERED XLSX — real Task → Service → SQLite → ExcelExporter ─────────


class TestFilteredXlsx:
    """FILTERED scope through the real headless task chain (no Qt boundary)."""

    def test_filtered_xlsx_real_task_chain_produces_openable_workbook(
        self, tmp_path,
    ) -> None:
        """ExportTask.execute → ExportService → real SQLite → ExcelExporter → file.

        The workbook is re-opened with openpyxl and its sheet / headers /
        data rows are asserted item by item, with the same leakage matrix.
        """
        from openpyxl import load_workbook

        context, seeded = _make_context(tmp_path)
        output_path = tmp_path / "filtered.xlsx"
        task = ExportTask(
            service=context.services.export,
            exporter=ExcelExporter(),
            output_path=str(output_path),
            scope=ExportScope.FILTERED,
            criteria=_PENDING_CRITERIA,
        )

        result = task.execute()

        assert result.startswith("Exported 5 rows")
        assert output_path.exists()

        workbook = load_workbook(str(output_path))
        sheet = workbook.active
        assert sheet is not None and sheet.title == "Export"
        assert [cell.value for cell in sheet[1]] == [
            "Person Name", "Department", "Note", "Photo Path", "Original Name",
            "Folder", "Captured At", "Match Confidence", "Match Status",
            "Archive Status", "Archive Target", "Archived At",
        ]
        assert sheet.max_row == 6  # header + 5 data rows

        values = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        sections = {
            "people": [], "photos": [], "matches": [], "archive": [],
        }
        for row in values:
            _, _, _, photo_path, _, _, _, _, match_status, archive_status, _, _ = [
                "" if cell is None else str(cell) for cell in row
            ]
            if match_status:
                sections["matches"].append(row)
            elif archive_status:
                sections["archive"].append(row)
            elif ".jpg" in photo_path:
                sections["photos"].append(row)
            else:
                sections["people"].append(row)

        # photos: main set = Photo A only.
        assert len(sections["photos"]) == 1
        assert "alice_portrait.jpg" in str(sections["photos"][0][3])
        assert str(sections["photos"][0][5]) == str(seeded["folder_id"])

        # matches: pending + approved of the main set, Photo A's id only.
        assert len(sections["matches"]) == 2
        statuses = {str(row[8]): float(row[7]) for row in sections["matches"]}
        assert statuses == {"pending": 0.87, "approved": 0.92}
        assert {str(row[3]) for row in sections["matches"]} == {str(seeded["photo_a_id"])}

        # people: Alice only (derived from the matches).
        assert [str(row[0]) for row in sections["people"]] == ["Alice"]

        # archive: Photo A's archived record only.
        assert len(sections["archive"]) == 1
        assert str(sections["archive"][0][9]) == "archived"
        assert str(sections["archive"][0][10]).endswith("alice_portrait.jpg")

        # Leakage: Photo B absent everywhere.
        flat = {str(cell) for row in values for cell in row}
        assert str(seeded["photo_b_id"]) not in flat
        assert "bob_candid.jpg" not in flat
        assert "Bob" not in flat
        assert "planned" not in flat and "rejected" not in flat


# ── C. CURRENT_BATCH → ValueError, no file, no silent fallback ───────────────


class TestCurrentBatchRejection:
    """CURRENT_BATCH is deferred (contract §2/D4): honest rejection everywhere."""

    def test_current_batch_rejected_by_real_task_with_contract_message(
        self, tmp_path,
    ) -> None:
        """Real ExportTask → ExportService raises the §2/D4 contract ValueError."""
        context, _ = _make_context(tmp_path)
        output_path = tmp_path / "current_batch.csv"
        task = ExportTask(
            service=context.services.export,
            exporter=CsvExporter(),
            output_path=str(output_path),
            scope=ExportScope.CURRENT_BATCH,
        )

        with pytest.raises(ValueError) as excinfo:
            task.execute()

        message = str(excinfo.value)
        assert "CURRENT_BATCH is deferred" in message
        assert "no batch persistence exists" in message
        assert "refusing to silently fall back to ALL" in message
        # No export happened: no file, and the ALL/FILTERED paths were not run.
        assert not output_path.exists()

    def test_current_batch_rejected_through_real_ui_chain(
        self, qtbot, tmp_path, monkeypatch,
    ) -> None:
        """UI-bypassed CURRENT_BATCH: real chain fails honestly, no fallback file.

        The real dialog disables this radio (Commit 3); here the boundary
        double bypasses it to prove the deeper Task → Service guard is
        reachable and produces no export file and no silent ALL/FILTERED run.
        """
        settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'current_batch.db'}")
        settings.ensure_runtime_directories()
        context = bootstrap_application(settings)
        _seed_scope_database(context.repositories)
        window = MainWindow(context)
        qtbot.addWidget(window)

        from PySide6.QtWidgets import QMessageBox

        warnings: list[tuple[str, str]] = []

        def _warning(parent, title: str, message: str) -> int:
            warnings.append((title, message))
            return 0

        monkeypatch.setattr(QMessageBox, "warning", staticmethod(_warning))
        output_path = tmp_path / "current_batch.csv"
        _stub_dialog(monkeypatch, output_path, ExportScope.CURRENT_BATCH, "csv")

        window._export_action.trigger()

        qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)
        assert window._status_label.text() == "export failed."
        # The contract error message reaches the user surface verbatim.
        assert warnings, "QMessageBox.warning must surface the TaskFailed event"
        assert warnings[0][0] == "Export Failed"
        assert "CURRENT_BATCH is deferred" in warnings[0][1]
        assert "refusing to silently fall back to ALL" in warnings[0][1]
        # No export file was produced — neither ALL, FILTERED, nor partial.
        assert not output_path.exists()


# ── D. FILTERED + criteria=None → ValueError through the real chain ─────────


class TestFilteredWithoutCriteriaRejection:
    """FILTERED without a criteria snapshot is rejected (contract §3/F3/F6)."""

    def test_filtered_without_criteria_rejected_by_real_task(
        self, tmp_path,
    ) -> None:
        """Real ExportTask(scope=FILTERED, criteria=None) raises the §3 ValueError."""
        context, _ = _make_context(tmp_path)
        output_path = tmp_path / "filtered_none.csv"
        task = ExportTask(
            service=context.services.export,
            exporter=CsvExporter(),
            output_path=str(output_path),
            scope=ExportScope.FILTERED,
            criteria=None,
        )

        with pytest.raises(ValueError) as excinfo:
            task.execute()

        message = str(excinfo.value)
        assert "FILTERED requires a PhotoSearchCriteria snapshot" in message
        assert "refusing to silently fall back to ALL" in message
        assert not output_path.exists()

    def test_filtered_without_criteria_rejected_through_real_ui_chain(
        self, qtbot, tmp_path, monkeypatch,
    ) -> None:
        """UI chain with no held filter: controller forwards None → honest failure.

        Not just the dialog disable (Commit 3, first invariant layer): the
        real MainWindow holds ``_current_criteria=None``, the handler forwards
        ``criteria=None`` per F5, and the Service's second invariant layer
        rejects — proving the guard is a real application-chain invariant.
        """
        settings = AppSettings(database_url=f"sqlite:///{tmp_path / 'filtered_none.db'}")
        settings.ensure_runtime_directories()
        context = bootstrap_application(settings)
        _seed_scope_database(context.repositories)
        window = MainWindow(context)
        qtbot.addWidget(window)
        assert window._current_criteria is None  # no filter ever set

        from PySide6.QtWidgets import QMessageBox

        warnings: list[tuple[str, str]] = []

        def _warning(parent, title: str, message: str) -> int:
            warnings.append((title, message))
            return 0

        monkeypatch.setattr(QMessageBox, "warning", staticmethod(_warning))
        output_path = tmp_path / "filtered_none.csv"
        created = _stub_dialog(monkeypatch, output_path, ExportScope.FILTERED, "csv")

        window._export_action.trigger()

        # The handler forwarded the None snapshot verbatim (F5).
        assert created and created[0].active_criteria is None
        qtbot.waitUntil(lambda: window._export_action.isEnabled(), timeout=_WAIT_TERMINAL_MS)
        assert window._status_label.text() == "export failed."
        assert warnings, "QMessageBox.warning must surface the TaskFailed event"
        assert "FILTERED requires a PhotoSearchCriteria snapshot" in warnings[0][1]
        assert not output_path.exists()


# ── E. ALL regression — historical behavior unchanged on real SQLite ────────


class TestAllScopeRegression:
    """ALL scope on real SQLite keeps the exact pre-FEATURE-004 behavior."""

    def test_all_scope_regression_on_real_sqlite_unchanged(self, tmp_path) -> None:
        """Real Task → Service → SQLite → CsvExporter → file, ALL semantics.

        Same seeded database as the FILTERED tests: ALL must still gather the
        full catalog with approved-only matches (pending excluded) — proving
        Commits 2/3 changed nothing about the ALL path (contract §6/AC-1).
        """
        context, seeded = _make_context(tmp_path)
        output_path = tmp_path / "all.csv"
        task = ExportTask(
            service=context.services.export,
            exporter=CsvExporter(),
            output_path=str(output_path),
            scope=ExportScope.ALL,
        )

        result = task.execute()

        assert result.startswith("Exported 8 rows")
        assert output_path.exists()
        headers, data_rows = _read_csv_rows(output_path)
        assert headers == _EXPECTED_HEADERS

        sections = _rows_by_section(data_rows)
        # people: full catalog — Alice AND Bob.
        assert {row[0] for row in sections["people"]} == {"Alice", "Bob"}
        # photos: full catalog — Photo A AND Photo B.
        assert {row[4] for row in sections["photos"]} == {
            "alice_portrait.jpg", "bob_candid.jpg",
        }
        # matches: approved-only per person (historical ALL semantics) —
        # Photo A's APPROVED (0.92) + Photo B's APPROVED (0.81); the PENDING
        # 0.87 result must NOT appear (this is the intentional ALL/FILTERED
        # asymmetry of contract §3/F4).
        assert len(sections["matches"]) == 2
        match_confidences = {row[7] for row in sections["matches"]}
        assert match_confidences == {"0.92", "0.81"}
        assert "0.87" not in match_confidences
        match_photo_ids = {row[3] for row in sections["matches"]}
        assert match_photo_ids == {
            str(seeded["photo_a_id"]), str(seeded["photo_b_id"]),
        }
        # archive: full history — both records, any status.
        assert len(sections["archive"]) == 2
        assert {row[9] for row in sections["archive"]} == {"archived", "planned"}
